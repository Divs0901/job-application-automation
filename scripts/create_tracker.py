"""
Creates the Job Application Tracker Excel file.
Run once to initialize your tracker.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'Job_Application_Tracker.xlsx')


def create_tracker():
    wb = Workbook()

    # ── Sheet 1: Applications ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"

    HEADER_COLOR   = "1F3864"   # dark navy
    ALT_ROW_COLOR  = "EEF2F7"   # light blue-grey
    ACCENT_COLOR   = "2E75B6"   # blue
    GREEN          = "70AD47"
    YELLOW         = "FFD966"
    RED_LIGHT      = "FF7070"

    # Column definitions: (header, width)
    columns = [
        ("#",              5),
        ("Date Applied",  14),
        ("Company",        22),
        ("Job Title",      28),
        ("Platform",       16),
        ("Location",       18),
        ("Salary Range",   16),
        ("Status",         18),
        ("Resume Version", 22),
        ("Resume Link",    30),
        ("Job URL",        35),
        ("Job Description",40),
        ("Cover Letter",   18),
        ("Contact Name",   20),
        ("Contact Email",  26),
        ("Follow-Up Date", 16),
        ("Interview Date", 16),
        ("Offer Amount",   16),
        ("Notes",          35),
    ]

    # Header row
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill   = PatternFill("solid", start_color=HEADER_COLOR)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )

    ws.row_dimensions[1].height = 36

    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # Pre-fill 100 rows with formulas and data-validation style formatting
    alt_fill   = PatternFill("solid", start_color=ALT_ROW_COLOR)
    white_fill = PatternFill("solid", start_color="FFFFFF")
    body_font  = Font(name="Arial", size=10)
    body_align = Alignment(vertical="center", wrap_text=False)

    STATUS_OPTIONS = ["To Apply", "Applied", "Phone Screen", "Interview",
                      "Technical Test", "Final Round", "Offer", "Rejected", "Withdrawn"]

    PLATFORM_OPTIONS = ["LinkedIn", "Indeed", "Glassdoor", "Company Site",
                        "Referral", "Recruiter", "Other"]

    for row in range(2, 102):
        fill = alt_fill if row % 2 == 0 else white_fill
        ws.row_dimensions[row].height = 20
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill      = fill
            cell.font      = body_font
            cell.alignment = body_align

        # Auto-number formula
        ws.cell(row=row, column=1).value = f'=IF(C{row}<>"",ROW()-1,"")'
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting for Status column (col 8 = H)
    green_fill  = PatternFill(start_color=GREEN,      end_color=GREEN,      fill_type="solid")
    yellow_fill = PatternFill(start_color=YELLOW,     end_color=YELLOW,     fill_type="solid")
    red_fill    = PatternFill(start_color=RED_LIGHT,  end_color=RED_LIGHT,  fill_type="solid")
    blue_fill   = PatternFill(start_color="9DC3E6",   end_color="9DC3E6",   fill_type="solid")

    status_range = "H2:H101"
    for value, fill in [
        ("Offer",         green_fill),
        ("Interview",     blue_fill),
        ("Final Round",   blue_fill),
        ("Applied",       yellow_fill),
        ("Phone Screen",  yellow_fill),
        ("Technical Test",yellow_fill),
        ("Rejected",      red_fill),
        ("Withdrawn",     red_fill),
    ]:
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill)
        )

    # ── Sheet 2: Dashboard ───────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False

    # Title
    dash.merge_cells("B2:H2")
    title_cell = dash["B2"]
    title_cell.value     = "📊 Job Application Dashboard"
    title_cell.font      = Font(name="Arial", bold=True, size=18, color=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    dash.row_dimensions[2].height = 40

    # Stats cards
    stats = [
        ("B4", "Total Applied",   "=COUNTIF(Applications!H2:H101,\"Applied\")+COUNTIF(Applications!H2:H101,\"Phone Screen\")+COUNTIF(Applications!H2:H101,\"Interview\")+COUNTIF(Applications!H2:H101,\"Technical Test\")+COUNTIF(Applications!H2:H101,\"Final Round\")+COUNTIF(Applications!H2:H101,\"Offer\")+COUNTIF(Applications!H2:H101,\"Rejected\")"),
        ("D4", "Interviews",      "=COUNTIF(Applications!H2:H101,\"Interview\")+COUNTIF(Applications!H2:H101,\"Final Round\")"),
        ("F4", "Offers",          "=COUNTIF(Applications!H2:H101,\"Offer\")"),
        ("H4", "Response Rate",   "=IFERROR(TEXT((COUNTIF(Applications!H2:H101,\"Phone Screen\")+COUNTIF(Applications!H2:H101,\"Interview\")+COUNTIF(Applications!H2:H101,\"Final Round\")+COUNTIF(Applications!H2:H101,\"Offer\"))/(COUNTIF(Applications!H2:H101,\"Applied\")+COUNTIF(Applications!H2:H101,\"Phone Screen\")+COUNTIF(Applications!H2:H101,\"Interview\")+COUNTIF(Applications!H2:H101,\"Final Round\")+COUNTIF(Applications!H2:H101,\"Offer\")+COUNTIF(Applications!H2:H101,\"Rejected\")),\"0%\"),\"0%\")"),
    ]

    card_colors = [ACCENT_COLOR, GREEN, "ED7D31", "7030A0"]
    for i, ((cell_ref, label, formula), color) in enumerate(zip(stats, card_colors)):
        # Label
        label_row = int(cell_ref[1:])
        col_letter = cell_ref[0]
        dash[cell_ref].value     = label
        dash[cell_ref].font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        dash[cell_ref].fill      = PatternFill("solid", start_color=color)
        dash[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        dash.row_dimensions[label_row].height = 24

        # Value cell (row below)
        val_ref = f"{col_letter}{label_row+1}"
        dash[val_ref].value     = formula
        dash[val_ref].font      = Font(name="Arial", bold=True, size=22, color=color)
        dash[val_ref].alignment = Alignment(horizontal="center", vertical="center")
        dash.row_dimensions[label_row+1].height = 36

        # Column width
        dash.column_dimensions[col_letter].width = 18

    # Platform breakdown table
    dash["B7"].value = "Applications by Platform"
    dash["B7"].font  = Font(name="Arial", bold=True, size=12, color=HEADER_COLOR)
    dash.row_dimensions[7].height = 24

    platforms = ["LinkedIn", "Indeed", "Glassdoor", "Company Site", "Referral", "Recruiter", "Other"]
    dash["B8"].value  = "Platform"
    dash["C8"].value  = "Count"
    dash["B8"].font   = dash["C8"].font = Font(name="Arial", bold=True, color="FFFFFF")
    dash["B8"].fill   = dash["C8"].fill = PatternFill("solid", start_color=HEADER_COLOR)

    for i, platform in enumerate(platforms, start=9):
        dash[f"B{i}"].value = platform
        dash[f"C{i}"].value = f'=COUNTIF(Applications!E2:E101,"{platform}")'
        dash[f"B{i}"].font  = dash[f"C{i}"].font = Font(name="Arial", size=10)
        if i % 2 == 0:
            dash[f"B{i}"].fill = dash[f"C{i}"].fill = PatternFill("solid", start_color=ALT_ROW_COLOR)

    dash.column_dimensions["B"].width = 20
    dash.column_dimensions["C"].width = 12

    # ── Sheet 3: Config ──────────────────────────────────────────────────────
    cfg = wb.create_sheet("Config")
    cfg["A1"].value = "Setting"
    cfg["B1"].value = "Value"
    cfg["A1"].font  = cfg["B1"].font = Font(bold=True, name="Arial")

    config_items = [
        ("Resume Folder",     "./resumes"),
        ("Base Resume File",  "base_resume.docx"),
        ("Anthropic API Key", "YOUR_ANTHROPIC_API_KEY"),
        ("LinkedIn Email",    "your@email.com"),
        ("LinkedIn Password", "your_password"),
        ("Indeed Email",      "your@email.com"),
        ("Indeed Password",   "your_password"),
        ("Glassdoor Email",   "your@email.com"),
        ("Glassdoor Password","your_password"),
        ("Default Location",  "Remote"),
        ("Phone Number",      "+1-555-000-0000"),
        ("Full Name",         "Your Name"),
    ]

    for i, (setting, value) in enumerate(config_items, start=2):
        cfg[f"A{i}"].value = setting
        cfg[f"B{i}"].value = value
        cfg[f"A{i}"].font  = Font(name="Arial", size=10, bold=True)
        cfg[f"B{i}"].font  = Font(name="Arial", size=10)
        if i % 2 == 0:
            cfg[f"A{i}"].fill = cfg[f"B{i}"].fill = PatternFill("solid", start_color=ALT_ROW_COLOR)

    cfg.column_dimensions["A"].width = 25
    cfg.column_dimensions["B"].width = 40
    cfg.protection.sheet = False

    wb.save(OUTPUT_PATH)
    print(f"✅ Tracker created: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_tracker()
